"""
rebuild_world0.py
=================
1. Parses Al-Bidaya wal-Nihaya (both volumes) into structured questions
2. Classifies them into creation topics for World 0
3. Moves Adam ﷺ–Nuh ﷺ questions from World 0 into World 1 (The Prophets)
4. Saves updated world0.json and world1.json
"""

import json, re, hashlib
from pathlib import Path
import openpyxl

BASE = Path('/home/ubuntu/seerah-quiz')
DATA = BASE / 'data'

# ── Creation topic classification ────────────────────────────────────────────
CREATION_TOPICS = [
    ("Beginning of Time · The Pen & The Tablet",
     ["pen","al-qalam","qalam","preserved tablet","lawh","lauh","write destinies","first thing allah created","first creation"]),
    ("Beginning of Time · The Throne & The Kursi",
     ["throne","arsh","kursi","footstool","carriers of the throne","bayt al-ma'mur","seven heavens","lowest heaven","second heaven","third heaven","fourth heaven","fifth heaven","sixth heaven","seventh heaven","heaven","heavens","firmament","sky","skies"]),
    ("Beginning of Time · Creation of the Earth",
     ["earth","land","mountains","water","seas","rivers","days of creation","six days","sunday","monday","tuesday","wednesday","thursday","friday","saturday","creation of earth","soil","dust","clay","ground","trees","plants","sustenance"]),
    ("Beginning of Time · Angels",
     ["angel","angels","mala'ikah","malaikah","jibril","gabriel","mikail","michael","israfil","azrael","izrail","munkar","nakir","kiraman","katibin","hamalat al-arsh","angel of death","wings of angels","light","created from light"]),
    ("Beginning of Time · Jinn & Iblis",
     ["jinn","iblis","shaytan","satan","devil","smokeless fire","fire","refused to prostrate","arrogance","iblis refused","jinn created","marij","smokeless","azazil","pride of iblis","disobedience of iblis"]),
    ("Beginning of Time · Creation of Adam ﷺ",
     ["adam","created from clay","created from dust","created from earth","breathed soul","prostrate to adam","sujud","taught names","al-asma","names of things","paradise","garden","forbidden tree","hawwa","eve","rib","companion for adam","before descent","in paradise","iblis swore","enmity","whispered to adam"]),
    ("Beginning of Time · Adam's Descent & Early Humanity",
     ["descent","descended","came down to earth","india","jeddah","habil","qabil","cain","abel","first murder","first idolatry","wadd","suwa","yaghuth","ya'uq","nasr","shayth","seth","early humanity","covenant","alast","am i not your lord","children of adam","progeny","black stone","leaves of paradise"]),
]

def classify_creation(text):
    text_lower = text.lower()
    best, best_score = None, 0
    for topic, keywords in CREATION_TOPICS:
        score = sum(1 for kw in keywords if kw in text_lower)
        if score > best_score:
            best_score = score
            best = topic
    return best if best_score > 0 else "Beginning of Time · Creation of the Earth"

def make_id(text):
    return 'bidaya_' + hashlib.md5(text.encode()).hexdigest()[:10]

def parse_bidaya(filepath):
    wb = openpyxl.load_workbook(filepath)
    questions = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else '' for c in row]
            text = ' '.join(c for c in cells if c).strip()
            if text:
                rows.append(text)

        i = 0
        while i < len(rows):
            line = rows[i]

            # Detect question lines: start with a number followed by . or )
            q_match = re.match(r'^(\d+)[\.\)]\s+(.+)', line)
            if not q_match:
                i += 1
                continue

            q_num = int(q_match.group(1))
            q_text = q_match.group(2).strip()

            # Collect multi-line question text
            j = i + 1
            while j < len(rows) and not re.match(r'^[a-dA-D][\.\)]\s', rows[j]) and not re.match(r'^\d+[\.\)]\s', rows[j]):
                next_line = rows[j]
                if re.match(r'^(Answer|ANSWER)\s*:', next_line):
                    break
                q_text += ' ' + next_line
                j += 1

            # Collect options a) b) c) d)
            opts = []
            correct = 0
            answer_line = ''

            while j < len(rows):
                opt_match = re.match(r'^([a-dA-D])[\.\)]\s+(.+)', rows[j])
                ans_match = re.match(r'^(Answer|ANSWER)\s*[:\-]\s*(.+)', rows[j])

                if opt_match:
                    opts.append(opt_match.group(2).strip())
                    j += 1
                elif ans_match:
                    answer_line = ans_match.group(2).strip()
                    j += 1
                    break
                elif rows[j].startswith('|') or re.match(r'^\d+\s*\|', rows[j]):
                    # Answer key table row — skip
                    j += 1
                    break
                else:
                    break

            if len(opts) < 2:
                i = j
                continue

            # Parse correct answer
            if answer_line:
                ans_match2 = re.match(r'^([a-dA-D])[\.\)]', answer_line)
                if ans_match2:
                    correct = ord(ans_match2.group(1).lower()) - ord('a')
                else:
                    # Try to find which option text matches
                    ans_clean = re.sub(r'^[a-dA-D][\.\)]\s*', '', answer_line).lower().strip()
                    for idx, opt in enumerate(opts):
                        if ans_clean and ans_clean[:30] in opt.lower():
                            correct = idx
                            break

            if len(opts) >= 2 and q_text:
                full_text = q_text + ' ' + ' '.join(opts)
                topic = classify_creation(full_text)
                questions.append({
                    'id': make_id(q_text),
                    'q': q_text,
                    'opts': opts[:4],
                    'correct': min(correct, len(opts)-1),
                    'topic': topic,
                    'difficulty': 'medium',
                    'source': 'Al-Bidaya wal-Nihaya'
                })

            i = j

    return questions

# ── Parse Al-Bidaya wal-Nihaya ───────────────────────────────────────────────
print("Parsing Al-Bidaya wal-Nihaya...")
bidaya_qs = parse_bidaya('/home/ubuntu/upload/pasted_file_UcMtht_Al-Bidayawal-Nihaya.xlsx')
print(f"  Parsed {len(bidaya_qs)} questions")

from collections import Counter
c = Counter(q['topic'] for q in bidaya_qs)
for t, n in sorted(c.items()):
    print(f"  {n:3d}  {t}")

# ── Separate creation vs Adam-as-prophet questions ───────────────────────────
# Questions about Adam's descent, children, early humanity go to World 0
# Questions about Adam as prophet (his prophethood story) go to World 1
WORLD0_TOPICS = {t for t, _ in CREATION_TOPICS}

world0_new = [q for q in bidaya_qs if q['topic'] in WORLD0_TOPICS]
# Adam's prophethood questions from bidaya also go to world1
adam_prophet_qs = []
for q in bidaya_qs:
    if q['topic'] not in WORLD0_TOPICS:
        q2 = dict(q)
        q2['topic'] = "The Prophets · Adam ﷺ"
        adam_prophet_qs.append(q2)

print(f"\nWorld 0 (creation): {len(world0_new)} questions")
print(f"World 1 (Adam prophet): {len(adam_prophet_qs)} questions")

# ── Load existing world data ─────────────────────────────────────────────────
with open(DATA / 'world0.json') as f:
    old_w0 = json.load(f)
with open(DATA / 'world1.json') as f:
    old_w1 = json.load(f)

# From old world0, move Adam/Idris/Nuh topics to world1
# Old world0 had "Beginning of Time · Adam ﷺ" etc.
w0_to_move = []
w0_keep = []
for q in old_w0:
    topic = q['topic']
    # Keep only pure creation topics
    if any(kw in topic for kw in ['Adam', 'Idris', 'Nuh', 'Prophets']):
        # Remap to The Prophets world
        q2 = dict(q)
        prophet = None
        for p in ['Adam ﷺ', 'Idris ﷺ', 'Nuh ﷺ']:
            if p in topic:
                prophet = p
                break
        if prophet:
            q2['topic'] = f"The Prophets · {prophet}"
        else:
            q2['topic'] = "The Prophets · Adam ﷺ"
        w0_to_move.append(q2)
    else:
        w0_keep.append(q)

print(f"\nFrom old world0: keeping {len(w0_keep)}, moving {len(w0_to_move)} to world1")

# ── Deduplicate by question text ─────────────────────────────────────────────
def dedup(questions):
    seen = set()
    out = []
    for q in questions:
        key = re.sub(r'\s+', ' ', q['q'].lower().strip())[:120]
        if key not in seen:
            seen.add(key)
            out.append(q)
    return out

# Build new world0: creation questions only
new_w0 = dedup(world0_new)
print(f"\nNew world0: {len(new_w0)} unique questions")

# Build new world1: existing + moved from w0 + adam_prophet from bidaya
new_w1 = dedup(old_w1 + w0_to_move + adam_prophet_qs)
print(f"New world1: {len(new_w1)} unique questions")

# Show world0 topic distribution
print("\n=== World 0 topics ===")
c0 = Counter(q['topic'] for q in new_w0)
for t, n in sorted(c0.items()):
    print(f"  {n:3d}  {t}")

print("\n=== World 1 Adam/Idris/Nuh topics ===")
c1 = Counter(q['topic'] for q in new_w1 if any(p in q['topic'] for p in ['Adam','Idris','Nuh']))
for t, n in sorted(c1.items()):
    print(f"  {n:3d}  {t}")

# ── Save ─────────────────────────────────────────────────────────────────────
with open(DATA / 'world0.json', 'w', encoding='utf-8') as f:
    json.dump(new_w0, f, ensure_ascii=False, indent=2)
with open(DATA / 'world1.json', 'w', encoding='utf-8') as f:
    json.dump(new_w1, f, ensure_ascii=False, indent=2)

print("\n✓ world0.json and world1.json saved.")
