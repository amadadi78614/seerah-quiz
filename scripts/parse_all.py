#!/usr/bin/env python3
"""
parse_all.py
============
Parses all Excel question files and outputs one JSON file per world
into the /data/ directory.

World structure (chronological):
  0 - beginning-of-time   (Adam, Idris, Nuh)
  1 - the-prophets        (Hud → Isa + prophethood theory)
  2 - pre-islam           (Arabia, Rome, Persia, world context)
  3 - seerah              (Life of the Prophet ﷺ)
  4 - the-sahaabah        (Companions + Men Around the Prophet)
  5 - post-islam          (Umayyads, Islamic Spain, Al-Bidaya later vols)

Usage:
  python3 scripts/parse_all.py [--input-dir /path/to/xlsx/files]
"""

import openpyxl, re, json, os, sys, glob, hashlib
from pathlib import Path

# ── Configuration ─────────────────────────────────────────────
BASE_DIR   = Path(__file__).parent.parent
DATA_DIR   = BASE_DIR / 'data'
DATA_DIR.mkdir(exist_ok=True)

# Map filename patterns to their known content type
FILE_ROUTING = {
    'preislam':         'preislam',
    'pre_islam':        'preislam',
    'pre-islam':        'preislam',
    'prophet':          'prophets',
    'bidaya':           'bidaya',
    'nihaya':           'bidaya',
    'seerah':           'seerah',
    'sealed':           'seerah',
    'nectar':           'seerah',
    'men':              'sahabah',
    'sahabah':          'sahabah',
    'companion':        'sahabah',
}

# Chronological topic ordering per world (used for sorting)
TOPIC_ORDER = {
    # World 0
    'Beginning of Time · Adam ﷺ':    10,
    'Beginning of Time · Idris ﷺ':   20,
    'Beginning of Time · Nuh ﷺ':     30,
    # World 1
    'The Prophets · Faith in Prophets':     5,
    'The Prophets · Who Are the Prophets':  6,
    'The Prophets · Hud ﷺ':          10,
    'The Prophets · Salih ﷺ':        20,
    'The Prophets · Ibrahim ﷺ':      30,
    "The Prophets · Isma'il ﷺ":      40,
    'The Prophets · Ishaq ﷺ':        50,
    'The Prophets · Yaqub ﷺ':        60,
    'The Prophets · Yusuf ﷺ':        70,
    'The Prophets · Ayyub ﷺ':        80,
    "The Prophets · Shu'aib ﷺ":      90,
    'The Prophets · Musa & Harun ﷺ': 100,
    'The Prophets · Hizqael ﷺ':      110,
    'The Prophets · Elisha ﷺ':       120,
    'The Prophets · Samuel ﷺ':       130,
    'The Prophets · Dawud ﷺ':        140,
    'The Prophets · Sulaiman ﷺ':     150,
    'The Prophets · Isaiah ﷺ':       160,
    'The Prophets · Jeremiah ﷺ':     170,
    'The Prophets · Daniel ﷺ':       180,
    'The Prophets · Uzair ﷺ':        190,
    'The Prophets · Yunus ﷺ':        200,
    'The Prophets · Zakariyya ﷺ':    210,
    'The Prophets · Yahya ﷺ':        220,
    'The Prophets · Isa ﷺ':          230,
    'The Prophets · Muhammad ﷺ':     240,
    'The Prophets · Their Mission':   250,
    'The Prophets · Their Qualities': 260,
    'The Prophets · Our Duty':        270,
    'The Prophets · Arabic Vocabulary': 280,
    'The Prophets · Mixed Review':    290,
    'The Prophets · Stories of the Prophets': 300,
    # World 2 (Pre-Islam)
    'Pre-Islam · World Civilisations': 10,
    'Pre-Islam · Arabia':              20,
    'Pre-Islam · Lineage & Ancestry':  30,
    "Pre-Islam · The Ka'bah":          40,
    'Pre-Islam · Religions of Arabia': 50,
    'Pre-Islam · Seekers of Truth':    60,
    'Pre-Islam · Early Life of the Prophet ﷺ': 70,
    'Pre-Islam · Khadijah & Marriages': 80,
    # World 3 (Seerah — chronological)
    'Seerah · Introduction to Seerah':  5,
    'Seerah · Pre-Islamic Context':    10,
    'Seerah · Early Life':             20,
    'Seerah · Before Prophethood':     30,
    'Seerah · The Revelation':         40,
    'Seerah · Foundations of Faith':   50,
    'Seerah · Persecution in Makkah':  60,
    'Seerah · Migration to Abyssinia': 70,
    'Seerah · Year of Grief':          80,
    "Seerah · At-Ta'if":               90,
    "Seerah · Al-Isra wal-Mi'raj":    100,
    'Seerah · Pledges of Aqabah':     110,
    'Seerah · The Hijrah':            120,
    'Seerah · Madinah':               130,
    'Seerah · Battle of Badr':        140,
    'Seerah · Battle of Uhud':        150,
    'Seerah · Battle of the Trench':  160,
    'Seerah · Hudaybiyyah & Later Events': 170,
    'Seerah · Detailed Studies':      180,
    'Seerah · Makkah':                 65,
    # World 4
    'The Sahaabah · The Companions':       10,
    'The Sahaabah · Stories of Conversion': 20,
    'The Sahaabah · Men Around the Prophet': 30,
    # World 5
    'Post-Islam · Umayyad Era':     10,
    'Post-Islam · Islamic Spain':   20,
    'Post-Islam · Islamic History': 30,
}

# ── Helpers ───────────────────────────────────────────────────

def get_rows(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                val = str(cell).strip()
                if val:
                    rows.append(val)
    return rows

def extract_questions(rows):
    questions = []
    i = 0
    current_section = 'General'
    while i < len(rows):
        line = rows[i]
        if (re.match(r'^Section \d+', line, re.I) or
            re.match(r'^Prophet ', line, re.I) or
            re.match(r'^CHAPTER ', line) or
            re.match(r'^MCQs Based', line, re.I) or
            re.match(r'^Stories of', line, re.I) or
            re.match(r'^ARABIC VOCABULARY', line, re.I) or
            re.match(r'^MIXED REVIEW', line, re.I)):
            current_section = line
            i += 1
            continue
        m = re.match(r'^(\d+)\.\s+(.+)', line)
        if m:
            qnum = int(m.group(1))
            qtext = re.sub(r'\s+', ' ', m.group(2)).strip()
            opts = []
            answer_letter = 'B'
            i += 1
            while i < len(rows):
                opt_line = rows[i]
                if re.match(r'^[A-Da-d]\)', opt_line):
                    letter = opt_line[0].upper()
                    text   = re.sub(r'\s+', ' ', opt_line[2:]).strip()
                    opts.append(f'{letter}. {text}')
                    i += 1
                elif opt_line.lower().startswith('answer:'):
                    m2 = re.match(r'[Aa]nswer:\s*([A-Da-d])', opt_line)
                    if m2:
                        answer_letter = m2.group(1).upper()
                    i += 1
                    break
                else:
                    break
            if len(opts) >= 2:
                questions.append({
                    'num': qnum, 'q': qtext, 'opts': opts,
                    'correct': {'A':0,'B':1,'C':2,'D':3}.get(answer_letter, 1),
                    'section': current_section
                })
        else:
            i += 1
    return questions

def q_key(q_text):
    """Dedup key: first 60 chars normalised."""
    return re.sub(r'\s+', ' ', q_text.lower().strip())[:60]

def q_id(q_text, prefix='q'):
    h = hashlib.md5(q_text.encode()).hexdigest()[:8]
    return f'{prefix}_{h}'

def difficulty_by_num(n, total):
    if n <= total * 0.33: return 'easy'
    if n <= total * 0.66: return 'medium'
    return 'hard'

# ── Topic assignment ──────────────────────────────────────────

def prophet_topic(section):
    s = section.lower()
    mapping = [
        (['adam'],                    (0, 'Beginning of Time · Adam \ufdfa')),
        (['idris','enoch'],           (0, 'Beginning of Time · Idris \ufdfa')),
        (['nuh','noah'],              (0, 'Beginning of Time · Nuh \ufdfa')),
        (['hud'],                     (1, 'The Prophets · Hud \ufdfa')),
        (['salih'],                   (1, 'The Prophets · Salih \ufdfa')),
        (['ibrahim','abraham'],       (1, 'The Prophets · Ibrahim \ufdfa')),
        (["isma'il",'ishmael'],       (1, "The Prophets · Isma'il \ufdfa")),
        (['ishaq','isaac'],           (1, 'The Prophets · Ishaq \ufdfa')),
        (['yaqub','jacob'],           (1, 'The Prophets · Yaqub \ufdfa')),
        (['lut','lot'],               (1, 'The Prophets · Lut \ufdfa')),
        (["shu'aib",'midian'],        (1, "The Prophets · Shu'aib \ufdfa")),
        (['yusuf','joseph'],          (1, 'The Prophets · Yusuf \ufdfa')),
        (['ayyub','job'],             (1, 'The Prophets · Ayyub \ufdfa')),
        (['dhul-kifl'],               (1, 'The Prophets · Dhul-Kifl \ufdfa')),
        (['yunus','jonah'],           (1, 'The Prophets · Yunus \ufdfa')),
        (['musa','moses','harun','aaron'], (1, 'The Prophets · Musa & Harun \ufdfa')),
        (['hizqael','ezekiel'],       (1, 'The Prophets · Hizqael \ufdfa')),
        (['elisha','elyas'],          (1, 'The Prophets · Elisha \ufdfa')),
        (['shammil','samuel'],        (1, 'The Prophets · Samuel \ufdfa')),
        (['dawud','david'],           (1, 'The Prophets · Dawud \ufdfa')),
        (['sulaiman','solomon'],      (1, 'The Prophets · Sulaiman \ufdfa')),
        (['isaiah','shia'],           (1, 'The Prophets · Isaiah \ufdfa')),
        (['aramaya','jeremiah'],      (1, 'The Prophets · Jeremiah \ufdfa')),
        (['daniel'],                  (1, 'The Prophets · Daniel \ufdfa')),
        (['uzair','ezra'],            (1, 'The Prophets · Uzair \ufdfa')),
        (['zakariyy','zechariah'],    (1, 'The Prophets · Zakariyya \ufdfa')),
        (['yahya','john'],            (1, 'The Prophets · Yahya \ufdfa')),
        (['isa','jesus'],             (1, 'The Prophets · Isa \ufdfa')),
        (['muhammad'],                (1, 'The Prophets · Muhammad \ufdfa')),
        (['fourth pillar','chapter 1'], (1, 'The Prophets · Faith in Prophets')),
        (['counting','chapter 2'],    (1, 'The Prophets · Who Are the Prophets')),
        (['mission','chapter 3'],     (1, 'The Prophets · Their Mission')),
        (['noble qualities','chapter 4'], (1, 'The Prophets · Their Qualities')),
        (['duty','chapter 5'],        (1, 'The Prophets · Our Duty')),
        (['arabic vocabulary'],       (1, 'The Prophets · Arabic Vocabulary')),
        (['mixed review'],            (1, 'The Prophets · Mixed Review')),
    ]
    for keywords, result in mapping:
        if any(kw in s for kw in keywords):
            return result
    return (1, 'The Prophets · Stories of the Prophets')

def preislam_topic(section, qnum):
    s = section.lower()
    if 'lineage' in s or 'ancestry' in s or 'section 1' in s:
        return (2, 'Pre-Islam · Lineage & Ancestry')
    if 'religious' in s or 'religion' in s or 'section 2' in s:
        return (2, 'Pre-Islam · Religions of Arabia')
    if 'early life' in s or 'elephant' in s or 'section 3' in s:
        return (2, 'Pre-Islam · Early Life of the Prophet \ufdfa')
    if 'khadijah' in s or 'marriage' in s or 'section 4' in s:
        return (2, 'Pre-Islam · Khadijah & Marriages')
    if "ka'bah" in s or 'kabah' in s or 'zamzam' in s or 'section 5' in s:
        return (2, "Pre-Islam · The Ka'bah")
    if 'seeker' in s or 'truth' in s or 'hunafa' in s or 'section 6' in s:
        return (2, 'Pre-Islam · Seekers of Truth')
    if 'rome' in s or 'byzantine' in s or 'persia' in s or 'india' in s or 'world' in s:
        return (2, 'Pre-Islam · World Civilisations')
    if 'arabia' in s or 'arab' in s or 'tribe' in s:
        return (2, 'Pre-Islam · Arabia')
    # Fallback by question number
    if qnum <= 20:  return (2, 'Pre-Islam · Lineage & Ancestry')
    if qnum <= 35:  return (2, 'Pre-Islam · Religions of Arabia')
    if qnum <= 55:  return (2, 'Pre-Islam · Early Life of the Prophet \ufdfa')
    if qnum <= 75:  return (2, 'Pre-Islam · Khadijah & Marriages')
    if qnum <= 84:  return (2, "Pre-Islam · The Ka'bah")
    return (2, 'Pre-Islam · Seekers of Truth')

def seerah_topic(section):
    s = section.lower()
    if 'introduction' in s or 'author' in s or 'importance' in s:
        return (3, 'Seerah · Introduction to Seerah')
    if 'background' in s or 'religious environment' in s or 'pre-islamic' in s or 'jahiliyyah' in s or 'context' in s:
        return (3, 'Seerah · Pre-Islamic Context')
    if 'elephant' in s or 'early life' in s or 'birth' in s:
        return (3, 'Seerah · Early Life')
    if 'truth' in s or 'glad-tiding' in s or 'before prophethood' in s:
        return (3, 'Seerah · Before Prophethood')
    if 'revelation' in s or 'secret' in s or 'dawah' in s:
        return (3, 'Seerah · The Revelation')
    if 'foundation' in s or 'aqeedah' in s or 'faith' in s or 'tawheed' in s or 'worship' in s:
        return (3, 'Seerah · Foundations of Faith')
    if 'open preaching' in s or 'opposition' in s or 'persecution' in s:
        return (3, 'Seerah · Persecution in Makkah')
    if 'abyssinia' in s:
        return (3, 'Seerah · Migration to Abyssinia')
    if 'year of grief' in s or 'siege' in s:
        return (3, 'Seerah · Year of Grief')
    if "ta'if" in s or 'jinn' in s:
        return (3, "Seerah · At-Ta'if")
    if 'isra' in s or "mi'raj" in s or 'night journey' in s:
        return (3, "Seerah · Al-Isra wal-Mi'raj")
    if 'aqabah' in s or 'search for a new home' in s or 'pledge' in s:
        return (3, 'Seerah · Pledges of Aqabah')
    if 'hijrah' in s or 'hijra' in s or 'migration to madinah' in s:
        return (3, 'Seerah · The Hijrah')
    if 'madinah' in s or 'pillars of the new' in s or 'merit' in s:
        return (3, 'Seerah · Madinah')
    if 'badr' in s:
        return (3, 'Seerah · Battle of Badr')
    if 'uhud' in s:
        return (3, 'Seerah · Battle of Uhud')
    if 'trench' in s or 'khandaq' in s or 'qurayzah' in s:
        return (3, 'Seerah · Battle of the Trench')
    if 'hudaybiyyah' in s or 'later events' in s or 'expulsion' in s:
        return (3, 'Seerah · Hudaybiyyah & Later Events')
    if 'fiqh' in s or 'minor' in s or 'anecdote' in s or 'expedition' in s:
        return (3, 'Seerah · Detailed Studies')
    return (3, 'Seerah · Makkah')

def sahabah_topic(section):
    s = section.lower()
    if 'embraced' in s or 'conversion' in s:
        return (4, 'The Sahaabah · Stories of Conversion')
    if 'men around' in s:
        return (4, 'The Sahaabah · Men Around the Prophet')
    return (4, 'The Sahaabah · The Companions')

def postislam_topic(section, qtext):
    q = qtext.lower()
    if 'spain' in q or 'tariq' in q or 'andalus' in q or 'gibraltar' in q:
        return (5, 'Post-Islam · Islamic Spain')
    if 'muawiya' in q or 'karbala' in q or 'husain' in q or 'yazid' in q:
        return (5, 'Post-Islam · Umayyad Era')
    return (5, 'Post-Islam · Islamic History')

def bidaya_topic(sheet_name, section, qnum, qtext):
    """Al-Bidaya wal-Nihaya spans creation through post-Islam."""
    sn = sheet_name.lower()
    s  = section.lower()
    q  = qtext.lower()
    # vol1 covers creation / prophets / pre-Islam
    if 'vol1' in sn:
        if qnum <= 60:
            return prophet_topic(section)  # creation / early prophets
        elif qnum <= 130:
            return (2, 'Pre-Islam · World Civilisations')
        else:
            return (3, 'Seerah · Pre-Islamic Context')
    # vol2+ covers Islamic history
    if 'conquest' in q or 'fath' in q or 'battle' in q or 'ghazwa' in q:
        return (3, 'Seerah · Hudaybiyyah & Later Events')
    if 'umayyad' in q or 'muawiya' in q or 'yazid' in q or 'karbala' in q:
        return (5, 'Post-Islam · Umayyad Era')
    return (5, 'Post-Islam · Islamic History')

# ── Main parse function ───────────────────────────────────────

def parse_file(filepath, file_type):
    """Parse an Excel file and return list of classified question dicts."""
    wb = openpyxl.load_workbook(filepath)
    all_entries = []

    for sheet_name in wb.sheetnames:
        rows = get_rows(wb[sheet_name])
        questions = extract_questions(rows)
        total = len(questions)

        for q in questions:
            diff = difficulty_by_num(q['num'], total)

            if file_type == 'prophets':
                sn = sheet_name.lower()
                # Sheet4 = pre-Islam world context
                if sheet_name in ('Sheet4',):
                    world, topic = (2, 'Pre-Islam · World Civilisations')
                # Sheet5 = Post-Islam Umayyads/Karbala
                elif sheet_name in ('Sheet5',):
                    world, topic = postislam_topic(q['section'], q['q'])
                # Sheet6 = Post-Islam Islamic Spain
                elif sheet_name in ('Sheet6',):
                    world, topic = (5, 'Post-Islam · Islamic Spain')
                # Sheets 7-11, 13 = Seerah
                elif sheet_name in ('Sheet7','Sheet8','Sheet9','Sheet10','Sheet11','Sheet13'):
                    world, topic = seerah_topic(q['section'])
                # Sheet12 = Pre-Islam Arabia
                elif sheet_name in ('Sheet12',):
                    world, topic = preislam_topic(q['section'], q['num'])
                # Sheet14 = Sahaabah
                elif sheet_name in ('Sheet14',):
                    world, topic = sahabah_topic(q['section'])
                # Sheets 1-3 = Prophets / Beginning of Time
                else:
                    world, topic = prophet_topic(q['section'])

            elif file_type == 'preislam':
                if q['num'] >= 67 and sheet_name in ('Sheet2','Sheet3'):
                    world, topic = seerah_topic(q['section'])
                else:
                    world, topic = preislam_topic(q['section'], q['num'])

            elif file_type == 'seerah':
                world, topic = seerah_topic(q['section'])

            elif file_type == 'sahabah':
                world, topic = sahabah_topic(q['section'])

            elif file_type == 'bidaya':
                result = bidaya_topic(sheet_name, q['section'], q['num'], q['q'])
                if isinstance(result, tuple) and len(result) == 2:
                    world, topic = result
                else:
                    world, topic = result  # already a tuple from prophet_topic

            else:
                world, topic = (2, 'Pre-Islam · Arabia')

            sort_key = TOPIC_ORDER.get(topic, 500)

            all_entries.append({
                'id':         q_id(q['q']),
                'difficulty': diff,
                'topic':      topic,
                'world':      world,
                'sort':       sort_key * 1000 + q['num'],
                'q':          q['q'],
                'opts':       q['opts'],
                'correct':    q['correct'],
                'explain':    '',
                'source':     os.path.basename(filepath),
            })

    return all_entries

# ── Collect all files ─────────────────────────────────────────

def detect_file_type(filepath):
    name = os.path.basename(filepath).lower()
    for pattern, ftype in FILE_ROUTING.items():
        if pattern in name:
            return ftype
    return 'unknown'

def find_input_files(input_dir):
    files = []
    for ext in ('*.xlsx', '*.xls'):
        files.extend(glob.glob(os.path.join(input_dir, ext)))
    return sorted(files)

# ── Main ──────────────────────────────────────────────────────

def main(input_dirs=None):
    if input_dirs is None:
        input_dirs = [
            str(BASE_DIR / 'new-questions'),
        ]

    all_entries = []
    for d in input_dirs:
        for fpath in find_input_files(d):
            ftype = detect_file_type(fpath)
            if ftype == 'unknown':
                print(f'  [SKIP] {os.path.basename(fpath)} — unrecognised file type')
                continue
            print(f'  [PARSE] {os.path.basename(fpath)} → {ftype}')
            try:
                entries = parse_file(fpath, ftype)
                all_entries.extend(entries)
                print(f'          {len(entries)} questions found')
            except Exception as e:
                print(f'  [ERROR] {e}')

    # Load existing data and merge
    worlds = {0:[], 1:[], 2:[], 3:[], 4:[], 5:[]}
    existing_keys = set()

    for wid in range(6):
        out_path = DATA_DIR / f'world{wid}.json'
        if out_path.exists():
            with open(out_path, 'r', encoding='utf-8') as f:
                existing = json.load(f)
            worlds[wid] = existing
            for q in existing:
                existing_keys.add(q_key(q['q']))

    # Add new unique questions
    added = 0
    for e in all_entries:
        k = q_key(e['q'])
        if k not in existing_keys:
            existing_keys.add(k)
            worlds[e['world']].append(e)
            added += 1

    # Sort each world chronologically
    for wid in range(6):
        worlds[wid].sort(key=lambda q: q.get('sort', 999999))

    # Write output
    for wid in range(6):
        out_path = DATA_DIR / f'world{wid}.json'
        with open(out_path, 'w', encoding='utf-8') as f:
            json.dump(worlds[wid], f, ensure_ascii=False, indent=2)
        print(f'  World {wid}: {len(worlds[wid])} questions → {out_path.name}')

    print(f'\nDone! {added} new questions added. Total: {sum(len(v) for v in worlds.values())}')
    return worlds

if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='Parse Excel question files into world JSON banks.')
    parser.add_argument('--input-dir', nargs='+', default=None,
                        help='Directory/directories containing .xlsx files (default: new-questions/)')
    args = parser.parse_args()
    print('Seerah Quiz — Question Parser')
    print('=' * 40)
    main(args.input_dir)
